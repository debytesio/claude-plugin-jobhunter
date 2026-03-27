#!/usr/bin/env python3
"""Process scraped job listings: deduplicate, score, calculate financial viability, export Excel.

Usage:
    python process_jobs.py --raw RAW_JSON --expectations EXPECTATIONS_JSON --config INI_FILE [options]

Options:
    --commute-data JSON Commute costs from get_commute_cost MCP tool
    --output-dir DIR    Output directory

Requires: _jh_core compiled C++ module for core algorithms.
"""

import argparse
import configparser
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# Add libs/ to path for platform-aware _jh_core package
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'libs'))

import _jh_core  # noqa: E402
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


# ── Operational defaults (overridden by INI) ──
DEFAULTS = {
    "hybrid_days_per_week": 2,
    "weeks_per_month": 4.33,
    "food_daily": 17.50,
}

# ── Commute cost tables (populated from country INI) ──
COMMUTE_COSTS = {}
LOCAL_TRANSPORT = {}
ACCOMMODATION = {}
OVERNIGHT_NEEDED = {}

# ── Currency formatting per country ──
CURRENCY_FORMATS = {
    "gb": {"symbol": "\u00a3", "code": "GBP", "fmt": "\u00a3{amount:,.0f}"},
    "fr": {"symbol": "\u20ac", "code": "EUR", "fmt": "\u20ac{amount:,.0f}"},
}


# ═══════════════════════════════════════════════════════════════════
# Data loading (I/O only — no algorithms)
# ═══════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════
# Resume text extraction (for .docx/.doc formats)
# ═══════════════════════════════════════════════════════════════════

def extract_resume_text(resume_path, output_dir=None):
    """Extract plain text from a resume file for LLM consumption.

    Supports: .tex, .pdf, .md, .txt (pass-through), .docx (python-docx), .doc (PowerShell COM).
    Returns the path to a readable text file.
    """
    path = Path(resume_path)
    ext = path.suffix.lower()
    if not path.exists():
        print(f"ERROR: Resume file not found: {resume_path}", file=sys.stderr)
        sys.exit(1)

    out_dir = Path(output_dir) if output_dir else path.parent

    if ext in (".txt", ".md", ".tex"):
        return str(path)

    if ext == ".pdf":
        try:
            import pymupdf
            doc = pymupdf.open(str(path))
            text = "\n\n".join(page.get_text() for page in doc)
            out_path = out_dir / f"{path.stem}-extracted.txt"
            out_path.write_text(text, encoding="utf-8")
            print(f"Extracted PDF text -> {out_path}")
            return str(out_path)
        except ImportError:
            print("ERROR: pymupdf not installed. Run: pip install pymupdf", file=sys.stderr)
            sys.exit(1)

    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(str(path))
            text = "\n".join(p.text for p in doc.paragraphs)
            out_path = out_dir / f"{path.stem}-extracted.txt"
            out_path.write_text(text, encoding="utf-8")
            print(f"Extracted .docx text -> {out_path}")
            return str(out_path)
        except ImportError:
            print("ERROR: python-docx not installed. Run: pip install python-docx", file=sys.stderr)
            sys.exit(1)

    if ext == ".doc":
        import subprocess
        out_path = out_dir / f"{path.stem}-extracted.txt"
        ps_script = (
            f'$w=New-Object -ComObject Word.Application;$w.Visible=$false;'
            f'$d=$w.Documents.Open("{path.resolve()}");'
            f'$d.SaveAs("{out_path.resolve()}",2);$d.Close();$w.Quit()'
        )
        try:
            result = subprocess.run(
                ['powershell', '-Command', ps_script],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode == 0 and out_path.exists():
                print(f"Extracted .doc via PowerShell COM -> {out_path}")
                return str(out_path)
            else:
                print(f"ERROR: PowerShell COM extraction failed: {result.stderr}", file=sys.stderr)
                sys.exit(1)
        except FileNotFoundError:
            print("ERROR: PowerShell not available. Cannot convert .doc files.", file=sys.stderr)
            sys.exit(1)
        except subprocess.TimeoutExpired:
            print("ERROR: PowerShell COM extraction timed out.", file=sys.stderr)
            sys.exit(1)

    print(f"ERROR: Unsupported resume format: {ext}", file=sys.stderr)
    print("Supported formats: .tex .pdf .docx .doc .txt .md", file=sys.stderr)
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════
# Company reputation cache management (I/O only)
# ═══════════════════════════════════════════════════════════════════

REPUTATION_CACHE_TTL_DAYS = 90


def _reputation_cache_path(country="gb"):
    """Return path to the persistent reputation cache file for a country."""
    return Path(__file__).resolve().parent.parent / "data" / f"company-reputation-cache-{country}.json"


def load_reputation_cache(country="gb"):
    """Load the persistent reputation cache from disk, filtering expired entries."""
    cache_path = _reputation_cache_path(country)
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            cache = json.load(f)
        from datetime import timedelta
        cutoff = (datetime.now() - timedelta(days=REPUTATION_CACHE_TTL_DAYS)).strftime("%Y-%m-%d")
        valid = {k: v for k, v in cache.items() if v.get("scraped_date", "") >= cutoff}
        return valid
    return {}


def save_reputation_cache(country, new_data, merge=True):
    """Save reputation data to the persistent cache file."""
    cache_path = _reputation_cache_path(country)
    existing = {}
    if merge and cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            existing = json.load(f)
    existing.update(new_data)
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)
    print(f"Saved {len(existing)} entries to reputation cache: {cache_path}")
    return str(cache_path)


def export_cache_snapshot(country, output_dir):
    """Export current valid cache entries to a snapshot file for the LLM agent."""
    cache = load_reputation_cache(country)
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    snapshot_path = Path(output_dir) / "reputation-cache-snapshot.json"
    with open(snapshot_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    print(f"Exported {len(cache)} cached reputation entries to {snapshot_path}")
    return str(snapshot_path)


def import_reputation_data(country, data_path):
    """Import newly scraped reputation data and merge into persistent cache."""
    with open(data_path, "r", encoding="utf-8") as f:
        new_data = json.load(f)
    cache_path = save_reputation_cache(country, new_data, merge=True)
    print(f"Merged {len(new_data)} new reputation entries into cache at {cache_path}")
    return cache_path


# ═══════════════════════════════════════════════════════════════════
# Configuration loading
# ═══════════════════════════════════════════════════════════════════

def load_config(ini_path, country="gb"):
    """Load shared INI + country-specific INI. Returns (dict, ConfigParser)."""
    cfg = configparser.ConfigParser()
    cfg.read(ini_path, encoding="utf-8")

    country_ini = Path(ini_path).parent / f"country-{country}.ini"
    if country_ini.exists():
        cfg.read(str(country_ini), encoding="utf-8")

    d = dict(DEFAULTS)
    d["country"] = country

    if cfg.has_section("tax"):
        for key in cfg.options("tax"):
            try:
                d[f"tax_{key}"] = cfg.getfloat("tax", key)
            except ValueError:
                d[f"tax_{key}"] = cfg.get("tax", key)

    if cfg.has_section("social"):
        for key in cfg.options("social"):
            try:
                d[f"social_{key}"] = cfg.getfloat("social", key)
            except ValueError:
                d[f"social_{key}"] = cfg.get("social", key)

    if cfg.has_section("commute_costs"):
        for k, v in cfg.items("commute_costs"):
            COMMUTE_COSTS[k] = float(v)
    if cfg.has_section("local_transport"):
        for k, v in cfg.items("local_transport"):
            LOCAL_TRANSPORT[k] = float(v)
    if cfg.has_section("accommodation_per_night"):
        for k, v in cfg.items("accommodation_per_night"):
            ACCOMMODATION[k] = float(v)
    if cfg.has_section("overnight_needed"):
        for k, v in cfg.items("overnight_needed"):
            OVERNIGHT_NEEDED[k] = v.strip() == "1"
    if cfg.has_section("food_daily"):
        d["food_daily"] = cfg.getfloat("food_daily", "default", fallback=d["food_daily"])

    if cfg.has_section("visa"):
        for key in cfg.options("visa"):
            try:
                d[f"visa_{key}"] = cfg.getfloat("visa", key)
            except ValueError:
                d[f"visa_{key}"] = cfg.get("visa", key)

    return d, cfg


def _load_commute_overrides(json_path):
    """Override commute globals with dynamic data from get_commute_cost MCP tool."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    results = (data if isinstance(data, list)
               else data.get("routes", data.get("results", [])))
    for route in results:
        if route.get("status") != "found":
            continue
        city = route["destination"].lower().strip()
        COMMUTE_COSTS[city] = route.get("return_fare", route.get("one_way_fare", 50) * 2)
        OVERNIGHT_NEEDED[city] = route.get("overnight_needed", False)

    print(f"Loaded commute data for {len(results)} destinations from MCP")


def load_candidate_skills(expectations, cfg_parser=None):
    """Load candidate skills from expectations JSON, fallback to INI [candidate_skills]."""
    skills = set()
    skill_data = expectations.get("candidate", {}).get("skills", {})
    if skill_data:
        for category in skill_data.values():
            if isinstance(category, list):
                skills.update(s.strip().lower() for s in category if s.strip())
    elif cfg_parser and cfg_parser.has_section("candidate_skills"):
        for _, val in cfg_parser.items("candidate_skills"):
            skills.update(s.strip().lower() for s in val.split(";") if s.strip())
    return skills




def _build_role_keywords(expectations):
    """Build role keyword sets from expectations JSON.

    Returns dict: {role_title: {"primary": [...], "related": [...]}}
    """
    generic_words = {
        "senior", "junior", "lead", "principal", "staff", "head",
        "director", "manager", "engineer", "developer", "architect",
        "consultant", "analyst", "specialist", "associate", "intern",
        "team", "of", "the", "and", "a", "an", "in", "for",
    }
    role_kw = {}
    for role in expectations.get("target_roles", []):
        title = role["title"]
        primary = [title.lower().strip()]
        for kw in role.get("search_keywords", []):
            primary.append(kw.lower().strip())
        related = set()
        for kw in primary:
            for word in kw.split():
                if word not in generic_words and len(word) >= 2:
                    related.add(word)
        role_kw[title] = {"primary": primary, "related": sorted(related)}
    return role_kw


def get_min_salary_for_role(target_role, expectations):
    """Get minimum target salary for a role from expectations."""
    for role in expectations.get("target_roles", []):
        if role["title"] == target_role:
            return role.get("min_salary", 70000)
    return 70000


def parse_academic_grade_salary(salary_text, cfg_parser):
    """Parse 'Grade X' salary patterns from academic job listings."""
    import re
    if not salary_text or not cfg_parser or not cfg_parser.has_section("academic_salary_grades"):
        return None, None

    grade_match = re.search(r'grade\s*(\d+)(?:\s*/\s*(\d+))?', salary_text, re.IGNORECASE)
    if grade_match:
        grade_low = int(grade_match.group(1))
        grade_high = int(grade_match.group(2)) if grade_match.group(2) else grade_low
        try:
            sal_min = cfg_parser.getfloat("academic_salary_grades", f"grade_{grade_low}_min")
            sal_max = cfg_parser.getfloat("academic_salary_grades", f"grade_{grade_high}_max")
            return int(sal_min), int(sal_max)
        except (configparser.NoOptionError, ValueError):
            pass

    if re.search(r'\bprofessor\b', salary_text, re.IGNORECASE):
        try:
            sal_min = cfg_parser.getfloat("academic_salary_grades", "professor_min")
            sal_max = cfg_parser.getfloat("academic_salary_grades", "professor_max")
            return int(sal_min), int(sal_max)
        except (configparser.NoOptionError, ValueError):
            pass

    return None, None


# ═══════════════════════════════════════════════════════════════════
# Excel helpers — data transforms
# ═══════════════════════════════════════════════════════════════════

PLATFORM_DOMAINS = {
    'reed': 'https://www.reed.co.uk',
    'indeed': 'https://uk.indeed.com',
    'linkedin': 'https://www.linkedin.com',
    'totaljobs': 'https://www.totaljobs.com',
    'cwjobs': 'https://www.cwjobs.co.uk',
    'cvlibrary': 'https://www.cv-library.co.uk',
    'adzuna': 'https://www.adzuna.co.uk',
}

PLATFORM_DISPLAY = {
    'reed': 'Reed', 'indeed': 'Indeed', 'linkedin': 'LinkedIn',
    'totaljobs': 'Totaljobs', 'cwjobs': 'CW Jobs', 'cvlibrary': 'CV-Library',
    'adzuna': 'Adzuna', 'indeed_fr': 'Indeed FR',
    'welcometothejungle': 'WTTJ', 'apec': 'APEC',
    'hellowork': 'HelloWork', 'lesjeudis': 'Les Jeudis',
}

_LOCATION_SUFFIXES = [
    ', England, United Kingdom', ', United Kingdom',
    ', Scotland, United Kingdom', ', Wales, United Kingdom',
    ', Lancashire', ', Cheshire', ', Berkshire', ', Yorkshire',
    ', Surrey', ', Essex', ', Kent', ', Hampshire', ', Oxfordshire',
    ', Cambridgeshire', ', Warwickshire', ', Hertfordshire',
]

_LOCATION_ALIASES = {
    'Greater Manchester': 'Manchester', 'Greater London': 'London',
    'City of London': 'London', 'Central London': 'London',
    'Manchester Area': 'Manchester', 'London Area': 'London',
}


def _normalize_location(loc):
    """Normalize location to clean city name."""
    if not loc:
        return ''
    for suffix in _LOCATION_SUFFIXES:
        loc = loc.replace(suffix, '')
    loc = re.sub(r'\s+[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}$', '', loc)
    loc = loc.strip().rstrip(',').strip()
    if loc in _LOCATION_ALIASES:
        loc = _LOCATION_ALIASES[loc]
    if loc.endswith(' Area'):
        loc = loc[:-5]
    return loc


def _normalize_url(url, platform):
    """Ensure URL is absolute."""
    if not url or url.startswith('http'):
        return url
    domain = PLATFORM_DOMAINS.get(platform, '')
    return f"{domain}{url}" if domain else url


def _format_salary(job, ccy):
    """Format salary as clean range or single value."""
    sal_min = job.get('salary_min', 0) or 0
    sal_max = job.get('salary_max', 0) or 0
    if not sal_min and not sal_max:
        return '\u2014'
    if sal_min == sal_max or not sal_max:
        return f"{ccy}{int(sal_min or sal_max):,}"
    return f"{ccy}{int(sal_min):,} \u2013 {ccy}{int(sal_max):,}"


def _humanize_job(job, ccy):
    """Transform raw pipeline values to human-readable format."""
    # Save raw platform before display transform
    job['_raw_platform'] = job.get('platform', '')
    # Booleans
    for key in ('viable', 'is_agency'):
        val = job.get(key)
        job[key] = 'Yes' if val in ('TRUE', True) else 'No'
    # Visa sponsor: UNKNOWN → No
    job['is_visa_sponsor'] = 'Yes' if job.get('is_visa_sponsor') == 'TRUE' else 'No'
    # Work mode
    mode = job.get('work_mode')
    if mode:
        job['work_mode'] = mode.capitalize()
    # Platform display name
    job['platform'] = PLATFORM_DISPLAY.get(
        job['_raw_platform'], job['_raw_platform'].title())
    # Location
    job['location'] = _normalize_location(job.get('location', ''))
    # URL
    job['job_url'] = _normalize_url(
        job.get('job_url', job.get('url', '')), job['_raw_platform'])
    # Salary display
    job['_salary_display'] = _format_salary(job, ccy)
    # Fix salary text encoding
    st = job.get('salary_text', '')
    if st and isinstance(st, str):
        job['salary_text'] = st.replace('\x80', ccy)


# ═══════════════════════════════════════════════════════════════════
# Excel output — styles and columns
# ═══════════════════════════════════════════════════════════════════

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="medium", color="2F5496"),
)

DATA_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
MUTED_FONT = Font(name="Calibri", size=10, color="999999")
MUTED_LINK_FONT = Font(name="Calibri", size=10, color="999999", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
NON_VIABLE_FILL = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
SPONSOR_YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AGENCY_YES_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
RATING_HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RATING_MED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
RATING_LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RATING_NA_FILL = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
VIABLE_YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
VIABLE_NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

SCORE_GREEN = Font(name="Calibri", size=10, bold=True, color="1F7A1F")
SCORE_ORANGE = Font(name="Calibri", size=10, color="CC7A00")
SCORE_RED = Font(name="Calibri", size=10, color="CC0000")

# Standard columns (21 — customer-facing)
STANDARD_COLUMNS = [
    ("rank", "#", 4),
    ("job_title", "Job Title", 45),
    ("company", "Company", 25),
    ("location", "Location", 18),
    ("work_mode", "Work Mode", 10),
    ("_salary_display", "Salary", 18),
    ("composite_score", "Score", 8),
    ("match_score", "Match", 8),
    ("financial_score", "Financial", 8),
    ("gross_annual", "Gross Annual", 13),
    ("net_monthly", "Net Monthly", 12),
    ("commute_monthly", "Commute/mo", 11),
    ("net_after_commute", "Take Home", 12),
    ("viable", "Viable", 8),
    ("is_visa_sponsor", "Visa Sponsor", 11),
    ("is_agency", "Agency", 8),
    ("company_rating", "Rating", 7),
    ("rating_reviews", "Reviews", 8),
    ("platform", "Platform", 10),
    ("posted_date", "Posted", 10),
    ("notes", "Notes", 20),
]

# Score Breakdown columns (extended — Pro/Power)
BREAKDOWN_COLUMNS = [
    ("rank", "#", 4),
    ("job_title", "Job Title", 40),
    ("company", "Company", 22),
    ("location", "Location", 16),
    ("composite_score", "Score", 8),
    ("match_score", "Match", 8),
    ("role_match", "Role", 7),
    ("skill_match", "Skill", 7),
    ("requirements_match", "Reqs", 7),
    ("experience_match", "Exp", 7),
    ("seniority_match", "Senior", 7),
    ("salary_match", "Salary", 7),
    ("location_priority", "Location", 7),
    ("sponsor_match", "Sponsor", 7),
    ("financial_score", "Financial", 8),
    ("salary_min", "Sal. Min", 10),
    ("salary_max", "Sal. Max", 10),
    ("salary_text", "Salary Text", 20),
    ("requirements_coverage", "Req. Coverage", 12),
    ("tech_stack_matched", "Tech Match", 12),
    ("yoe_required", "YoE Required", 10),
    ("viable", "Viable", 8),
]

_CURRENCY_KEYS = {"gross_annual", "net_monthly", "commute_monthly",
                  "net_after_commute", "salary_min", "salary_max"}
_SCORE_KEYS = {"match_score", "composite_score", "financial_score",
               "role_match", "skill_match", "requirements_match",
               "experience_match", "seniority_match", "salary_match",
               "location_priority", "sponsor_match"}
_CENTER_KEYS = _CURRENCY_KEYS | _SCORE_KEYS | {
    "rank", "work_mode", "viable", "is_visa_sponsor", "is_agency",
    "company_rating", "rating_reviews", "platform", "posted_date"}


def _score_comment(job):
    """Build cell comment for score breakdown."""
    lines = []
    for label, key in [("Role", "role_match"), ("Skill", "skill_match"),
                       ("Reqs", "requirements_match"), ("Exp", "experience_match"),
                       ("Senior", "seniority_match"), ("Salary", "salary_match"),
                       ("Location", "location_priority"), ("Sponsor", "sponsor_match")]:
        val = job.get(key)
        lines.append(f"{label}: {val:.0f}" if isinstance(val, (int, float)) else f"{label}: \u2014")
    return f"{lines[0]} | {lines[1]} | {lines[2]}\n{lines[3]} | {lines[4]} | {lines[5]}\n{lines[6]} | {lines[7]}"


def _rating_comment(job):
    """Build cell comment for rating source. Returns None if no source."""
    source = job.get('rating_source')
    if not source or source == 'not_found':
        return None
    reviews = int(job.get('rating_reviews', 0) or 0)
    if reviews:
        return f"Source: {source.title()} ({reviews:,} reviews)"
    return f"Source: {source.title()}"


def write_data_sheet(ws, jobs, columns=None, exclude_cols=None, ccy_fmt='\u00a3#,##0'):
    """Write a formatted data sheet with job listings."""
    cols = columns or STANDARD_COLUMNS
    if exclude_cols:
        cols = [(k, l, w) for k, l, w in cols if k not in exclude_cols]
    is_standard = columns is None or columns is STANDARD_COLUMNS

    # Headers
    for ci, (key, label, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    for ri, job in enumerate(jobs, 2):
        is_alt = ri % 2 == 0
        is_viable = job.get('viable') == 'Yes'
        row_fill = NON_VIABLE_FILL if not is_viable else (ALT_ROW_FILL if is_alt else None)
        row_font = MUTED_FONT if not is_viable else DATA_FONT

        for ci, (key, label, width) in enumerate(cols, 1):
            val = job.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = row_font
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            # Job Title as hyperlink
            if key == "job_title":
                url = job.get('job_url', '')
                if url:
                    cell.hyperlink = url
                    cell.font = MUTED_LINK_FONT if not is_viable else LINK_FONT

            # Currency
            elif key in _CURRENCY_KEYS and isinstance(val, (int, float)):
                cell.number_format = ccy_fmt
                if key == "commute_monthly" and not val:
                    cell.value = '\u2014'

            # Scores
            elif key in _SCORE_KEYS and isinstance(val, (int, float)):
                cell.number_format = '0'

            # Composite score color
            if key == "composite_score" and isinstance(val, (int, float)) and is_viable:
                if val >= 80:
                    cell.font = SCORE_GREEN
                elif val < 40:
                    cell.font = SCORE_RED
                elif val < 60:
                    cell.font = SCORE_ORANGE

            # Status cell fills (override row fill)
            if is_viable:
                if key == "viable":
                    cell.fill = VIABLE_YES_FILL
                elif key == "is_visa_sponsor" and val == 'Yes':
                    cell.fill = SPONSOR_YES_FILL
                elif key == "is_agency" and val == 'Yes':
                    cell.fill = AGENCY_YES_FILL
                elif key == "company_rating":
                    if isinstance(val, (int, float)):
                        cell.number_format = '0.0'
                        if val >= 4.0:
                            cell.fill = RATING_HIGH_FILL
                        elif val >= 3.0:
                            cell.fill = RATING_MED_FILL
                        else:
                            cell.fill = RATING_LOW_FILL
                    else:
                        cell.fill = RATING_NA_FILL
                        cell.value = "N/A"
            elif key == "viable":
                cell.fill = VIABLE_NO_FILL

            # Reviews
            if key == "rating_reviews" and isinstance(val, (int, float)) and val:
                cell.number_format = '#,##0'

            # Center alignment
            if key in _CENTER_KEYS:
                cell.alignment = Alignment(horizontal="center")

        # Cell comments (standard sheets only)
        if is_standard:
            score_ci = next((i for i, (k, _, _) in enumerate(cols, 1)
                             if k == "composite_score"), None)
            if score_ci:
                ws.cell(row=ri, column=score_ci).comment = Comment(
                    _score_comment(job), "DEB Cloud")
            rating_ci = next((i for i, (k, _, _) in enumerate(cols, 1)
                              if k == "company_rating"), None)
            if rating_ci:
                rc = _rating_comment(job)
                if rc:
                    ws.cell(row=ri, column=rating_ci).comment = Comment(rc, "DEB Cloud")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(jobs) + 1}"


def write_summary_sheet(ws, stats, ccy='\u00a3'):
    """Write structured summary sheet with visual sections."""
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    title_font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    subtitle_font = Font(name="Calibri", size=11, color="666666")
    section_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    label_font = Font(name="Calibri", size=10, color="555555")
    value_font = Font(name="Calibri", bold=True, size=10)
    metric_font = Font(name="Calibri", bold=True, size=18, color="2F5496")
    metric_label = Font(name="Calibri", size=9, color="888888")
    center = Alignment(horizontal="center")

    def _section(row, text, span=4):
        for c in range(1, span + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = section_fill
            cell.font = section_font
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        return row + 1

    def _kv(row, label, value):
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value).font = value_font
        return row + 1

    row = 1
    ws.cell(row=row, column=1, value="Job Search Report").font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {stats['search_date']}").font = subtitle_font
    row += 2

    # Overview
    row = _section(row, "OVERVIEW")
    row += 1
    for col, (val, lbl) in enumerate([
        (stats['total_matched'], 'Jobs Scored'),
        (stats['total_viable'], 'Financially Viable'),
        (stats.get('total_companies', 0), 'Companies'),
    ], 1):
        ws.cell(row=row, column=col, value=val).font = metric_font
        ws.cell(row=row, column=col).alignment = center
        ws.cell(row=row + 1, column=col, value=lbl).font = metric_label
        ws.cell(row=row + 1, column=col).alignment = center
    row += 3

    # Financial target
    row = _section(row, "FINANCIAL TARGET")
    row = _kv(row, "Current Net Monthly", stats.get('current_net_monthly', 'N/A'))
    row = _kv(row, "Target Net Monthly", stats.get('target_net_monthly', 'N/A'))
    row = _kv(row, "Improvement Threshold", f"+{stats.get('improvement_pct', 10)}%")
    row = _kv(row, "Visa Sponsorship Required", stats.get('requires_visa', 'N/A'))
    row += 1

    # By location
    if stats.get('city_counts'):
        row = _section(row, "BY LOCATION")
        ws.cell(row=row, column=1, value="City").font = label_font
        ws.cell(row=row, column=2, value="Jobs").font = label_font
        ws.cell(row=row, column=3, value="Viable").font = label_font
        row += 1
        viable_by_city = stats.get('viable_by_city', {})
        for city, count in sorted(stats['city_counts'].items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=city).font = value_font
            ws.cell(row=row, column=2, value=count).font = value_font
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=3, value=viable_by_city.get(city, 0)).font = value_font
            ws.cell(row=row, column=3).alignment = center
            row += 1
        row += 1

    # Company insights
    if stats.get('reputation_stats'):
        row = _section(row, "COMPANY INSIGHTS")
        rep = stats['reputation_stats']
        row = _kv(row, "Companies Rated", rep.get('with_rating', 0))
        row = _kv(row, "Average Rating", f"{rep.get('avg_rating', 0):.1f} / 5.0")
        row = _kv(row, "Agencies Detected", stats.get('total_agencies', 0))
        row = _kv(row, "Visa Sponsors", stats.get('total_sponsors', 0))
        row += 1

    # Platforms
    if stats.get('platform_counts'):
        row = _section(row, "PLATFORMS")
        for plat, count in sorted(stats['platform_counts'].items(), key=lambda x: -x[1]):
            row = _kv(row, plat, f"{count} jobs")
        row += 1

    # Per role
    if stats.get('role_counts'):
        row = _section(row, "BY ROLE")
        for role, count in stats['role_counts'].items():
            row = _kv(row, role, f"{count} jobs")


def write_help_sheet(ws):
    """Write the Guide sheet explaining columns and color coding."""
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80

    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    section_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    name_font = Font(name="Calibri", bold=True, size=10)
    desc_font = Font(name="Calibri", size=10)

    def _section(row, text):
        for c in range(1, 3):
            cell = ws.cell(row=row, column=c)
            cell.fill = section_fill
            cell.font = section_font
        ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        return row + 1

    row = 1
    ws.cell(row=row, column=1, value="Report Guide").font = title_font
    row += 2

    row = _section(row, "COLUMN DESCRIPTIONS")
    for name, desc in [
        ("Job Title", "Job posting title. Click to open the original listing."),
        ("Company", "Hiring company name."),
        ("Location", "City where the job is based."),
        ("Work Mode", "Remote, Hybrid, or Onsite."),
        ("Salary", "Salary range or single value. '\u2014' if not listed."),
        ("Score", "Overall match (0\u2013100). = Match \u00d7 60% + Financial \u00d7 40%. "
         "Hover for the full 8-dimension breakdown."),
        ("Match", "How well the job fits your profile: role, skills, seniority, "
         "salary expectations, location, and visa sponsorship."),
        ("Financial", "Financial improvement vs your current take-home, "
         "after tax, NI, and commute costs. 100 = maximum improvement."),
        ("Gross Annual", "Estimated annual salary (midpoint if range)."),
        ("Net Monthly", "Monthly take-home after income tax and NI."),
        ("Commute/mo", "Monthly commute cost (rail fare \u00d7 working days). "
         "'\u2014' if remote or local."),
        ("Take Home", "Net Monthly minus Commute/mo. Your actual monthly income."),
        ("Viable", "Whether Take Home meets your target (current net + improvement %)."),
        ("Visa Sponsor", "Whether the company is a registered UKVI Skilled Worker sponsor."),
        ("Agency", "Whether the company is a recruitment agency (not the direct employer)."),
        ("Rating", "Employee rating (x.x / 5.0). Hover for source and review count."),
        ("Reviews", "Number of employee reviews the rating is based on."),
        ("Platform", "Job board where the listing was found."),
        ("Posted", "Date the job was originally posted."),
    ]:
        ws.cell(row=row, column=1, value=name).font = name_font
        ws.cell(row=row, column=2, value=desc).font = desc_font
        row += 1

    row += 1
    row = _section(row, "COLOR CODING")
    for name, desc in [
        ("Light red row", "Job is not financially viable (below your target)."),
        ("Green Score", "Excellent match (score \u2265 80)."),
        ("Orange Score", "Moderate match (score 40\u201359)."),
        ("Red Score", "Weak match (score < 40)."),
        ("Green Viable/Sponsor", "Yes \u2014 meets criteria."),
        ("Amber Agency", "Company is a recruitment agency."),
        ("Green Rating", "Good employer rating (\u2265 4.0)."),
        ("Amber Rating", "Average rating (3.0\u20133.9)."),
        ("Red Rating", "Low rating (< 3.0)."),
        ("Grey Rating", "No rating data available."),
    ]:
        ws.cell(row=row, column=1, value=name).font = name_font
        ws.cell(row=row, column=2, value=desc).font = desc_font
        row += 1

    row += 1
    row = _section(row, "SCORING METHOD")
    for line in [
        "Score = Match \u00d7 60% + Financial \u00d7 40%",
        "",
        "Match is based on 8 dimensions (hover over Score for breakdown):",
        "  \u2022 Skill Match (22%) \u2014 tech stack vs your skills",
        "  \u2022 Requirements Match (18%) \u2014 hard requirements vs your experience",
        "  \u2022 Role Match (15%) \u2014 job title vs your target roles",
        "  \u2022 Experience Match (12%) \u2014 years of experience required",
        "  \u2022 Seniority Match (10%) \u2014 seniority level alignment",
        "  \u2022 Salary Match (10%) \u2014 salary vs your expectations",
        "  \u2022 Location Priority (8%) \u2014 preferred vs acceptable cities",
        "  \u2022 Sponsor Match (5%) \u2014 UKVI sponsorship availability",
    ]:
        ws.cell(row=row, column=1, value=line).font = desc_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1


# ═══════════════════════════════════════════════════════════════════
# Excel generation
# ═══════════════════════════════════════════════════════════════════

def _generate_excel(scored_jobs, expectations, output_dir, expectations_path="",
                    total_scraped=None, total_dedup=None):
    """Generate Excel workbook from scored+financial-computed job list."""
    results_per_group = expectations.get("preferences", {}).get("results_per_group", 20)
    current_net = expectations.get("current_situation", {}).get("net_monthly_take_home", 0)
    improvement_pct = expectations.get("preferences", {}).get("improvement_threshold_pct", 10)
    requires_visa = expectations.get("visa", {}).get("requires_visa", False)
    country = expectations.get("country", "gb")
    ccy = CURRENCY_FORMATS.get(country, CURRENCY_FORMATS["gb"])["symbol"]
    ccy_fmt = f"{ccy}#,##0"

    # Humanize all jobs for display
    for job in scored_jobs:
        _humanize_job(job, ccy)

    def cap_per_city(jobs_list, limit):
        city_counts = {}
        result = []
        for j in jobs_list:
            city_key = j["location"].lower().split(",")[0].strip()
            city_counts[city_key] = city_counts.get(city_key, 0) + 1
            if city_counts[city_key] <= limit:
                result.append(j)
        return result

    role_order = [r["title"] for r in sorted(expectations.get("target_roles", []),
                                             key=lambda r: r.get("priority", 99))]
    role_jobs = {}
    for role in role_order:
        role_list = [j for j in scored_jobs if j.get("target_role") == role]
        role_list.sort(key=lambda x: (0 if x.get("region_group") == "P1" else 1,
                                      -x.get("composite_score", 0)))
        role_list = cap_per_city(role_list, results_per_group)
        for i, j in enumerate(role_list, 1):
            j["rank"] = i
        role_jobs[role] = role_list

    now = datetime.now()
    filename = f"jobs-{now.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    output_path = Path(output_dir) / filename

    wb = Workbook()
    first_sheet = True
    for role in role_order:
        if first_sheet:
            ws = wb.active
            ws.title = role
            first_sheet = False
        else:
            ws = wb.create_sheet(role)
        write_data_sheet(ws, role_jobs.get(role, []),
                         exclude_cols={"target_role"}, ccy_fmt=ccy_fmt)

    # All Results
    ws_all = wb.create_sheet("All Results")
    all_sorted = sorted([dict(j) for j in scored_jobs],
                        key=lambda x: -x.get("composite_score", 0))
    for i, j in enumerate(all_sorted, 1):
        j["rank"] = i
    write_data_sheet(ws_all, all_sorted, ccy_fmt=ccy_fmt)

    # Score Breakdown (extended)
    ws_breakdown = wb.create_sheet("Score Breakdown")
    write_data_sheet(ws_breakdown, all_sorted, columns=BREAKDOWN_COLUMNS, ccy_fmt=ccy_fmt)

    # Summary stats
    total_viable = sum(1 for j in scored_jobs if j.get("viable") == "Yes")
    target_net = round(current_net * (1 + improvement_pct / 100), 2) if current_net else "N/A"

    city_counts = {}
    viable_by_city = {}
    for j in scored_jobs:
        city = j.get("location", "").split(",")[0].strip()
        if city:
            city_counts[city] = city_counts.get(city, 0) + 1
            if j.get("viable") == "Yes":
                viable_by_city[city] = viable_by_city.get(city, 0) + 1

    platform_counts = {}
    for j in scored_jobs:
        p = j.get("platform", "")
        if p:
            platform_counts[p] = platform_counts.get(p, 0) + 1

    companies = set(j.get("company", "") for j in scored_jobs if j.get("company"))
    total_agencies = sum(1 for j in scored_jobs if j.get("is_agency") == "Yes")
    total_sponsors = sum(1 for j in scored_jobs if j.get("is_visa_sponsor") == "Yes")
    role_counts = {role: len(jlist) for role, jlist in role_jobs.items()}
    rated_jobs = [j for j in scored_jobs if isinstance(j.get("company_rating"), (int, float))]

    summary_stats = {
        "search_date": now.strftime("%Y-%m-%d %H:%M:%S"),
        "total_scraped": total_scraped if total_scraped is not None else len(scored_jobs),
        "total_dedup": total_dedup if total_dedup is not None else len(scored_jobs),
        "total_matched": len(scored_jobs),
        "total_viable": total_viable,
        "total_companies": len(companies),
        "total_agencies": total_agencies,
        "total_sponsors": total_sponsors,
        "current_net_monthly": f"{ccy}{current_net:,.2f}" if current_net else "N/A",
        "improvement_pct": improvement_pct,
        "target_net_monthly": f"{ccy}{target_net:,.2f}" if isinstance(target_net, float) else target_net,
        "requires_visa": "Yes" if requires_visa else "No",
        "city_counts": city_counts,
        "viable_by_city": viable_by_city,
        "platform_counts": platform_counts,
        "role_counts": role_counts,
    }
    if rated_jobs:
        sources = [j.get("rating_source", "") for j in rated_jobs if j.get("rating_source")]
        summary_stats["reputation_stats"] = {
            "with_rating": len(rated_jobs),
            "without_rating": len(scored_jobs) - len(rated_jobs),
            "avg_rating": sum(j["company_rating"] for j in rated_jobs) / len(rated_jobs),
            "primary_source": max(set(sources), key=sources.count) if sources else "N/A",
        }

    ws_summary = wb.create_sheet("Summary")
    write_summary_sheet(ws_summary, summary_stats, ccy)

    # Guide
    ws_guide = wb.create_sheet("Guide")
    write_help_sheet(ws_guide)

    wb.save(output_path)
    return str(output_path), role_jobs, role_order, role_counts


# ═══════════════════════════════════════════════════════════════════
# Pipeline stages
# ═══════════════════════════════════════════════════════════════════

def process_dedup(raw_path, output_dir):
    """Stage 1: Load raw jobs, deduplicate, save checkpoint-dedup.json."""
    with open(raw_path, "r", encoding="utf-8") as f:
        jobs = json.load(f)
    print(f"Loaded {len(jobs)} raw jobs")

    unique = list(_jh_core.dedup_jobs(jobs))
    print(f"After dedup: {len(unique)} unique jobs")

    checkpoint_dir = Path(output_dir)
    dedup_path = checkpoint_dir / "checkpoint-dedup.json"
    with open(dedup_path, "w", encoding="utf-8") as f:
        json.dump(unique, f, ensure_ascii=False, indent=2)
    print(f"Saved dedup checkpoint: {dedup_path}")
    return str(dedup_path)


def process_filter(dedup_path, expectations_path, output_dir,
                   threshold=40):
    """Stage 1.5: First-layer filter — drop low-relevance jobs before enrichment.

    Uses listing_score (from parser) + salary floor.
    Free — no API calls. Reduces enrichment volume by 30-40%.
    """
    with open(dedup_path, "r", encoding="utf-8") as f:
        jobs = json.load(f)
    with open(expectations_path, "r", encoding="utf-8") as f:
        expectations = json.load(f)

    min_salary = 0
    for role in expectations.get("target_roles", []):
        sal = role.get("min_salary", 0)
        if sal:
            min_salary = min(min_salary, sal) if min_salary else sal

    salary_floor = int(min_salary * 0.7) if min_salary else 0

    filtered = []
    dropped = {"low_score": 0, "salary": 0}
    for job in jobs:
        if job.get("listing_score", 100) < threshold:
            dropped["low_score"] += 1
            continue
        sal_max = job.get("salary_max", 0)
        if salary_floor and sal_max and sal_max < salary_floor:
            dropped["salary"] += 1
            continue
        filtered.append(job)

    checkpoint_dir = Path(output_dir)
    filter_path = checkpoint_dir / "checkpoint-filtered.json"
    with open(filter_path, "w", encoding="utf-8") as f:
        json.dump(filtered, f, ensure_ascii=False, indent=2)

    print(f"Filter: {len(jobs)} -> {len(filtered)} "
          f"(dropped: score={dropped['low_score']}, "
          f"salary={dropped['salary']})")
    return str(filter_path)


def process_enrich_prep(filtered_path, output_dir, max_jobs=-1):
    """Prepare enrichment UUID list from filtered jobs.

    Sorts by listing_score desc, takes top N (-1 = all), outputs
    a JSON array of UUIDs for the enrich MCP tool.
    """
    with open(filtered_path) as f:
        jobs = json.load(f)

    jobs.sort(key=lambda j: j.get('listing_score', 0), reverse=True)
    top = jobs[:max_jobs] if max_jobs > 0 else jobs
    uuids = [j['uuid'] for j in top if j.get('uuid')]

    out_path = Path(output_dir) / 'enrich-payload.json'
    with open(out_path, 'w') as f:
        json.dump(uuids, f)

    print(f"Enrich prep: {len(jobs)} filtered -> {len(uuids)} to enrich")
    return str(out_path)


def process_enrich_merge(filtered_path, enrich_path, output_dir):
    """Merge enrichment JD data back into filtered jobs.

    Sets jd_fetched/jd_status and JD fields on matched jobs.
    Outputs checkpoint-enriched.json.
    """
    with open(filtered_path, encoding='utf-8') as f:
        jobs = json.load(f)
    with open(enrich_path, encoding='utf-8') as f:
        jds = json.load(f)

    jd_map = {jd['uuid']: jd for jd in jds if jd.get('uuid')}
    enriched = 0
    for job in jobs:
        jd = jd_map.get(job.get('uuid'))
        if jd:
            job['jd_fetched'] = True
            job['jd_status'] = 'enriched'
            for field in ('responsibilities', 'requirements_hard',
                          'requirements_soft', 'tech_stack',
                          'seniority_signals', 'yoe_required',
                          'education'):
                job[field] = jd.get(field)
            enriched += 1
        else:
            job['jd_fetched'] = False
            job['jd_status'] = 'unavailable'

    out_path = Path(output_dir) / 'checkpoint-enriched.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(jobs, f, ensure_ascii=False, indent=2)

    print(f"Enrich merge: {enriched}/{len(jobs)} enriched, "
          f"{len(jobs) - enriched} unavailable")
    return str(out_path)


def process_extract_companies(scored_path, output_dir):
    """Extract unique company names from scored jobs.

    Outputs a JSON array of company names for reputation lookup.
    """
    with open(scored_path, encoding='utf-8') as f:
        jobs = json.load(f)

    companies = sorted(set(
        j.get('company', '') for j in jobs
        if j.get('company') and j.get('match_score', 0) >= 30))

    out_path = Path(output_dir) / 'companies.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(companies, f, ensure_ascii=False)

    print(f"Extract companies: {len(companies)} unique "
          f"(from {len(jobs)} scored jobs)")
    return str(out_path)


def process_reputation_merge(scored_path, reputation_path, output_dir):
    """Merge company check data (reputation + UKVI + agency) into jobs.

    Adds company_rating, rating_reviews, rating_source,
    is_visa_sponsor, sponsor_route, is_agency to each job.
    """
    with open(scored_path, encoding='utf-8') as f:
        jobs = json.load(f)
    with open(reputation_path, encoding='utf-8') as f:
        checks = json.load(f)

    check_map = {}
    for c in checks:
        name = c.get('company', '')
        if name:
            check_map[name] = c

    rated = sponsors = agencies = 0
    for job in jobs:
        check = check_map.get(job.get('company', ''), {})
        # Reputation
        if check.get('reputation_status') == 'found':
            job['company_rating'] = check.get('rating')
            job['rating_reviews'] = check.get('review_count', 0)
            job['rating_source'] = check.get('source', 'not_found')
            rated += 1
        else:
            job['company_rating'] = None
            job['rating_reviews'] = 0
            job['rating_source'] = 'not_found'
        # UKVI sponsor
        is_sponsor = check.get('is_sponsor', False)
        job['is_visa_sponsor'] = 'TRUE' if is_sponsor else 'UNKNOWN'
        job['sponsor_route'] = check.get('sponsor_route', '')
        if is_sponsor:
            sponsors += 1
        # Agency
        is_agency = check.get('is_agency', False)
        job['is_agency'] = 'TRUE' if is_agency else 'FALSE'
        if is_agency:
            agencies += 1

    out_path = Path(output_dir) / 'checkpoint-scored.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(jobs, f, ensure_ascii=False, indent=2)

    print(f"Company checks merge: {rated}/{len(jobs)} rated, "
          f"{sponsors} sponsors, {agencies} agencies")
    return str(out_path)


def process_excel(
        scored_path, expectations_path, config_path,
        output_dir=None, commute_data_path=None):
    """Stage 3: Financial calc + visa + Excel from pre-scored jobs."""
    with open(expectations_path, "r", encoding="utf-8") as f:
        expectations = json.load(f)

    country = expectations.get("country", "gb")
    cfg, cfg_parser = load_config(config_path, country=country)

    if commute_data_path and Path(commute_data_path).exists():
        _load_commute_overrides(commute_data_path)
    ccy = CURRENCY_FORMATS.get(country, CURRENCY_FORMATS["gb"])["symbol"]

    # Agency detection is now handled by company checks MCP (is_agency field)
    # Empty detector for C++ signature compat
    agency_detector = _jh_core.AgencyDetector(set(), [], [])

    current_net = expectations.get("current_situation", {}).get("net_monthly_take_home", 0)
    improvement_pct = expectations.get("preferences", {}).get("improvement_threshold_pct", 10)
    requires_visa = expectations.get("visa", {}).get("requires_visa", False)
    hybrid_days = expectations.get("preferences", {}).get("hybrid_days_per_week", 2)
    cfg["hybrid_days_per_week"] = hybrid_days

    # UKVI sponsor data is now set directly on jobs by company checks MCP tool
    # (is_visa_sponsor, sponsor_route fields). Empty index for C++ signature compat.
    ukvi_index = _jh_core.UkviIndex({})

    with open(scored_path, "r", encoding="utf-8") as f:
        scored_data = json.load(f)
    # Support both flat list and {scored_jobs: [...]} wrapper
    scored_input = (scored_data.get("scored_jobs", scored_data)
                    if isinstance(scored_data, dict) else scored_data)
    print(f"Loaded {len(scored_input)} pre-scored jobs")

    # Fix encoding artifacts and dates
    _fix_text_encoding(scored_input)
    _fix_posted_dates(scored_input)

    # Ensure target_role is assigned
    role_keywords = _build_role_keywords(expectations)
    default_role = expectations["target_roles"][0]["title"]
    for job in scored_input:
        if not job.get("target_role"):
            # Use C++ scoring to find best role match
            best_role = default_role
            best_score = 0
            title = job.get("title", job.get("job_title", ""))
            for role_title in role_keywords:
                sc = _jh_core.score_role_match(title, role_title, role_keywords)
                if sc > best_score:
                    best_score = sc
                    best_role = role_title
            job["target_role"] = best_role

    # Tag contract roles before processing
    for job in scored_input:
        if job.get("contract_type") == "contract":
            job["notes"] = (
                job.get("notes", "") + " Day-rate contract").strip()

    # Assign region_group from P1/P2 cities
    locs = expectations.get('locations', {})
    p1_cities = {c.lower() for c in
                 locs.get('p1', {}).get('cities', [])}
    p2_cities = {c.lower() for c in
                 locs.get('p2', {}).get('cities', [])}
    for job in scored_input:
        if not job.get('region_group'):
            loc = job.get('location', '').lower()
            job['region_group'] = ('P1' if any(
                c in loc for c in p1_cities) else 'P2')

    # Compute requirements_coverage + tech_stack_matched
    candidate_skills = [s.lower() for s in
                        expectations.get('candidate_skills', [])]
    for job in scored_input:
        reqs = job.get('requirements_hard', [])
        if isinstance(reqs, list) and reqs and candidate_skills:
            matched = sum(1 for r in reqs
                          if any(s in r.lower()
                                 for s in candidate_skills))
            job['requirements_coverage'] = (
                f"{int(matched / len(reqs) * 100)}%")
        tech = job.get('tech_stack', [])
        if isinstance(tech, list) and tech and candidate_skills:
            matched = [t for t in tech
                       if t.lower() in candidate_skills]
            job['tech_stack_matched'] = (
                ', '.join(matched) if matched else '')

    # Sanitize types for C++ compatibility
    for job in scored_input:
        if job.get('salary_min') is None:
            job['salary_min'] = 0
        if job.get('salary_max') is None:
            job['salary_max'] = 0
        # Remove list/dict fields C++ doesn't expect
        for key in list(job.keys()):
            if isinstance(job[key], (list, dict)):
                del job[key]

    # Process all jobs via C++ multi-threaded batch processor
    final_jobs = list(_jh_core.process_batch(
        scored_input, cfg, ukvi_index, agency_detector,
        current_net, improvement_pct, country,
        COMMUTE_COSTS, LOCAL_TRANSPORT, ACCOMMODATION, OVERNIGHT_NEEDED))
    print(f"After filtering (match>=30): {len(final_jobs)} jobs")

    out_dir = output_dir or str(Path(expectations_path).parent)

    # Save final checkpoint
    checkpoint_dir = Path(out_dir)
    with open(checkpoint_dir / "checkpoint-final.json", "w", encoding="utf-8") as f:
        json.dump(final_jobs, f, ensure_ascii=False, indent=2)

    # Generate Excel
    xlsx_path, role_jobs, role_order, role_counts = _generate_excel(
        final_jobs, expectations, out_dir,
        expectations_path=expectations_path,
        total_scraped=len(scored_input),
        total_dedup=len(scored_input),
    )

    # Jobs are now humanized (Yes/No) after _generate_excel
    total_viable = sum(1 for j in final_jobs if j.get("viable") == "Yes")
    print(f"\nExcel written to: {xlsx_path}")
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Pre-scored input:   {len(scored_input)}")
    print(f"After filtering:    {len(final_jobs)}")
    print(f"Financially viable: {total_viable}")
    for role in role_order:
        print(f"  {role:30s}: {role_counts.get(role, 0)} jobs")

    for role in role_order:
        top_list = role_jobs.get(role, [])[:10]
        if not top_list:
            continue
        print(f"\n{'-'*60}")
        print(f"TOP 10 {role}")
        print(f"{'-'*60}")
        for j in top_list:
            sal = f"{ccy} {j['gross_annual']:,}" if j["gross_annual"] else "Unlisted"
            sponsor = "[Y]" if j.get("is_visa_sponsor") == "Yes" else "[?]"
            viable_str = "[Y]" if j.get("viable") == "Yes" else "[X]"
            net_str = f"Net: {ccy} {j['net_after_commute']:,.0f}/mo" if j.get("net_after_commute") else ""
            line = (f"  #{j['rank']:2d} [{j['composite_score']:5.1f}] {j['job_title'][:40]:40s}"
                    f" | {j['company'][:18]:18s} | {j['location']:10s} | {sal:>12s}"
                    f" | {sponsor} | {viable_str} | {net_str}")
            print(line)

    return xlsx_path


def process_jobs(
        raw_path, expectations_path, config_path,
        output_dir=None, commute_data_path=None):
    """Full pipeline (--stage all): dedup -> score -> financial -> Excel."""
    with open(expectations_path, "r", encoding="utf-8") as f:
        expectations = json.load(f)

    country = expectations.get("country", "gb")
    cfg, cfg_parser = load_config(config_path, country=country)

    if commute_data_path and Path(commute_data_path).exists():
        _load_commute_overrides(commute_data_path)
    ccy = CURRENCY_FORMATS.get(country, CURRENCY_FORMATS["gb"])["symbol"]

    # Load candidate skills
    candidate_skills = load_candidate_skills(expectations, cfg_parser)

    # Agency detection handled by company checks MCP
    agency_detector = _jh_core.AgencyDetector(set(), [], [])

    sector = expectations.get("sector", "industry")
    current_net = expectations.get("current_situation", {}).get("net_monthly_take_home", 0)
    improvement_pct = expectations.get("preferences", {}).get("improvement_threshold_pct", 10)
    results_per_group = expectations.get("preferences", {}).get("results_per_group", 20)
    requires_visa = expectations.get("visa", {}).get("requires_visa", False)
    hybrid_days = expectations.get("preferences", {}).get("hybrid_days_per_week", 2)
    cfg["hybrid_days_per_week"] = hybrid_days

    # UKVI sponsor data is now set directly on jobs by company checks MCP tool
    ukvi_index = _jh_core.UkviIndex({})

    # Load raw jobs
    with open(raw_path, "r", encoding="utf-8") as f:
        jobs = json.load(f)
    print(f"Loaded {len(jobs)} raw jobs")

    # Deduplicate via C++
    unique_jobs = list(_jh_core.dedup_jobs(jobs))
    print(f"After dedup: {len(unique_jobs)} unique jobs")

    # Save dedup checkpoint
    checkpoint_dir = Path(output_dir or Path(expectations_path).parent)
    dedup_path = checkpoint_dir / "checkpoint-dedup.json"
    with open(dedup_path, "w", encoding="utf-8") as f:
        json.dump(unique_jobs, f, ensure_ascii=False, indent=2)
    print(f"Saved dedup checkpoint: {dedup_path}")

    # Build role keywords and score each job via C++
    role_keywords = _build_role_keywords(expectations)
    default_role = expectations["target_roles"][0]["title"]

    scored_jobs = []
    for job in unique_jobs:
        title = job.get("title", "")

        # Assign target role via C++ scoring
        if job.get("target_role"):
            target_role = job["target_role"]
        else:
            best_role = default_role
            best_score = 0
            for role_title in role_keywords:
                sc = _jh_core.score_role_match(title, role_title, role_keywords)
                if sc > best_score:
                    best_score = sc
                    best_role = role_title
            target_role = best_role
        min_salary = get_min_salary_for_role(target_role, expectations)

        # Parse academic grade salary if needed
        if sector == "academia" and job.get("salary_unlisted", True):
            grade_min, grade_max = parse_academic_grade_salary(
                job.get("salary_text", ""), cfg_parser)
            if grade_min is not None:
                job["salary_min"] = grade_min
                job["salary_max"] = grade_max
                job["salary_unlisted"] = False
                job["salary_text"] = job.get("salary_text", "") + f" (Grade->{grade_min:,}-{grade_max:,})"

        # Score via C++
        role_sc = _jh_core.score_role_match(title, target_role, role_keywords)
        skill_sc = _jh_core.score_skill_match(title, job.get("description", ""), candidate_skills)
        seniority_sc = _jh_core.score_seniority(title, sector)
        salary_sc = _jh_core.score_salary(
            job.get("salary_min"), job.get("salary_max"),
            job.get("salary_unlisted", True), min_salary)
        location_sc = _jh_core.score_location(
            job.get("location", ""), job.get("work_mode", "hybrid"),
            job.get("region_group", "P1"))
        sponsor_sc = _jh_core.score_sponsor(
            job.get("company", ""), ukvi_index) if requires_visa else 0

        if sector == "academia":
            if requires_visa:
                match_score = (role_sc * 0.30 + skill_sc * 0.25 + seniority_sc * 0.10
                               + salary_sc * 0.10 + location_sc * 0.10 + sponsor_sc * 0.15)
            else:
                match_score = (role_sc * 0.35 + skill_sc * 0.30 + seniority_sc * 0.10
                               + salary_sc * 0.15 + location_sc * 0.10)
        else:
            if requires_visa:
                match_score = (role_sc * 0.25 + skill_sc * 0.20 + seniority_sc * 0.10
                               + salary_sc * 0.20 + location_sc * 0.10 + sponsor_sc * 0.15)
            else:
                match_score = (role_sc * 0.30 + skill_sc * 0.25 + seniority_sc * 0.15
                               + salary_sc * 0.20 + location_sc * 0.10)

        if match_score < 30:
            continue

        job["target_role"] = target_role
        job["match_score"] = round(match_score, 1)
        job["role_match"] = role_sc
        job["skill_match"] = skill_sc
        job["seniority_match"] = seniority_sc
        job["salary_match"] = salary_sc
        job["location_priority"] = location_sc
        job["sponsor_match"] = sponsor_sc
        scored_jobs.append(job)

    print(f"After scoring (match>=30): {len(scored_jobs)} jobs")

    # Save scored checkpoint
    scored_path = checkpoint_dir / "checkpoint-scored.json"
    with open(scored_path, "w", encoding="utf-8") as f:
        json.dump(scored_jobs, f, ensure_ascii=False, indent=2)
    print(f"Saved scored checkpoint: {scored_path}")

    # Process financial viability via C++ multi-threaded batch
    final_jobs = list(_jh_core.process_batch(
        scored_jobs, cfg, ukvi_index, agency_detector,
        current_net, improvement_pct, country,
        COMMUTE_COSTS, LOCAL_TRANSPORT, ACCOMMODATION, OVERNIGHT_NEEDED))

    # Sort, rank, generate Excel
    def cap_per_city(jobs_list, limit):
        city_counts = {}
        result = []
        for j in jobs_list:
            city_key = j["location"].lower().split(",")[0].strip()
            city_counts[city_key] = city_counts.get(city_key, 0) + 1
            if city_counts[city_key] <= limit:
                result.append(j)
        return result

    out_dir = str(checkpoint_dir)
    xlsx_path, role_jobs, role_order, role_counts = _generate_excel(
        final_jobs, expectations, out_dir,
        expectations_path=expectations_path,
        total_scraped=len(jobs),
        total_dedup=len(unique_jobs),
    )

    # Jobs are now humanized (Yes/No) after _generate_excel
    total_viable = sum(1 for j in final_jobs if j.get("viable") == "Yes")
    print(f"\nExcel written to: {xlsx_path}")
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Total scraped:      {len(jobs)}")
    print(f"After dedup:        {len(unique_jobs)}")
    print(f"After scoring:      {len(final_jobs)}")
    print(f"Financially viable: {total_viable}")
    for role in role_order:
        print(f"  {role:30s}: {role_counts.get(role, 0)} jobs")

    for role in role_order:
        top_list = role_jobs.get(role, [])[:10]
        if not top_list:
            continue
        print(f"\n{'-'*60}")
        print(f"TOP 10 {role}")
        print(f"{'-'*60}")
        for j in top_list:
            sal = f"{ccy} {j['gross_annual']:,}" if j["gross_annual"] else "Unlisted"
            sponsor = "[Y]" if j.get("is_visa_sponsor") == "Yes" else "[?]"
            viable_str = "[Y]" if j.get("viable") == "Yes" else "[X]"
            net_str = f"Net: {ccy} {j['net_after_commute']:,.0f}/mo" if j.get("net_after_commute") else ""
            line = (f"  #{j['rank']:2d} [{j['composite_score']:5.1f}] {j['job_title'][:40]:40s}"
                    f" | {j['company'][:18]:18s} | {j['location']:10s} | {sal:>12s}"
                    f" | {sponsor} | {viable_str} | {net_str}")
            print(line)

    return str(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Process scraped job listings into Excel")
    parser.add_argument("--stage",
                        choices=["dedup", "filter", "enrich-prep", "enrich-merge",
                                 "extract-companies", "reputation-merge",
                                 "excel", "all", "extract-resume",
                                 "load-reputation-cache", "save-reputation-cache"],
                        default="all",
                        help="Pipeline stage: dedup, filter, excel, all, extract-resume, "
                             "load-reputation-cache, save-reputation-cache")
    parser.add_argument("--raw", help="Path to raw scraped jobs JSON (required for dedup/all)")
    parser.add_argument("--scored", help="Path to pre-scored jobs JSON (required for excel stage)")
    parser.add_argument("--resume", help="Path to resume file (for extract-resume stage)")
    parser.add_argument("--reputation-data", help="Path to new reputation data JSON (for save-reputation-cache)")
    parser.add_argument("--expectations", help="Path to expectations JSON")
    parser.add_argument("--config", help="Path to INI config")
    parser.add_argument("--commute-data", help="Path to commute cost JSON from get_commute_cost MCP tool")
    parser.add_argument("--dedup", help="Path to dedup checkpoint JSON (required for filter stage)")
    parser.add_argument("--filtered", help="Path to filtered checkpoint JSON (required for enrich-prep/enrich-merge)")
    parser.add_argument("--enrich-data", help="Path to enrichment results JSON (required for enrich-merge)")
    parser.add_argument("--max-enrich", type=int, default=-1,
                        help="Max jobs to enrich (-1 = all)")
    parser.add_argument("--filter-threshold", type=int, default=40,
                        help="Listing score threshold for filter stage (default: 40)")
    parser.add_argument("--output-dir", help="Output directory (defaults to expectations JSON dir)")
    args = parser.parse_args()

    output_dir = args.output_dir or (str(Path(args.expectations).parent) if args.expectations else ".")

    if args.stage == "extract-resume":
        if not args.resume:
            parser.error("--resume is required for --stage extract-resume")
        result_path = extract_resume_text(args.resume, output_dir)
        print(f"RESUME_TEXT_PATH={result_path}")

    elif args.stage == "load-reputation-cache":
        if not args.expectations:
            parser.error("--expectations is required for --stage load-reputation-cache")
        with open(args.expectations, "r", encoding="utf-8") as f:
            expectations = json.load(f)
        country = expectations.get("country", "gb")
        export_cache_snapshot(country, output_dir)

    elif args.stage == "save-reputation-cache":
        if not args.reputation_data:
            parser.error("--reputation-data is required for --stage save-reputation-cache")
        if not args.expectations:
            parser.error("--expectations is required for --stage save-reputation-cache")
        with open(args.expectations, "r", encoding="utf-8") as f:
            expectations = json.load(f)
        country = expectations.get("country", "gb")
        import_reputation_data(country, args.reputation_data)

    elif args.stage == "dedup":
        if not args.raw:
            parser.error("--raw is required for --stage dedup")
        process_dedup(args.raw, output_dir)

    elif args.stage == "filter":
        if not args.dedup:
            parser.error("--dedup is required for --stage filter")
        if not args.expectations:
            parser.error("--expectations is required for --stage filter")
        process_filter(args.dedup, args.expectations, output_dir,
                       threshold=args.filter_threshold)

    elif args.stage == "enrich-prep":
        if not args.filtered:
            parser.error("--filtered is required for --stage enrich-prep")
        process_enrich_prep(args.filtered, output_dir,
                            max_jobs=args.max_enrich)

    elif args.stage == "enrich-merge":
        if not args.filtered:
            parser.error("--filtered is required for --stage enrich-merge")
        if not args.enrich_data:
            parser.error("--enrich-data is required for --stage enrich-merge")
        process_enrich_merge(args.filtered, args.enrich_data, output_dir)

    elif args.stage == "extract-companies":
        if not args.scored:
            parser.error("--scored is required for --stage extract-companies")
        process_extract_companies(args.scored, output_dir)

    elif args.stage == "reputation-merge":
        if not args.scored:
            parser.error("--scored is required for --stage reputation-merge")
        if not args.reputation_data:
            parser.error("--reputation-data is required for --stage reputation-merge")
        process_reputation_merge(args.scored, args.reputation_data, output_dir)

    elif args.stage == "excel":
        if not args.scored:
            parser.error("--scored is required for --stage excel")
        if not args.expectations or not args.config:
            parser.error("--expectations and --config are required for --stage excel")
        process_excel(args.scored, args.expectations, args.config,
                      output_dir=output_dir,
                      commute_data_path=args.commute_data)

    else:  # all
        if not args.raw:
            parser.error("--raw is required for --stage all")
        if not args.expectations or not args.config:
            parser.error("--expectations and --config are required for --stage all")
        process_jobs(args.raw, args.expectations, args.config,
                     output_dir=output_dir,
                     commute_data_path=args.commute_data)


if __name__ == "__main__":
    main()
